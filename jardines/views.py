from django.shortcuts import render, get_object_or_404, redirect
from django.http import JsonResponse, HttpResponse
from django.views.decorators.csrf import csrf_exempt
from django.contrib import messages

from django.db.models import Count, Q
from .models import Programa, Subprograma, Jardin, Sala, AsistenciaDocente
from users.decorators import rol_requerido
from django.contrib.auth.decorators import login_required
from users.models import Usuario
from datetime import date
import calendar
import csv
from django.utils import timezone
from django.views.generic import TemplateView
from alumnos.models import AsignacionSala, Alumno
import json
from django.db.models.functions import TruncMonth
import openpyxl
import os

@login_required
@rol_requerido("coordinador", "administrador")
def cargar_subprogramas(request):
    programa_id = request.GET.get("programa_id")
    subprogramas = Subprograma.objects.filter(programa_id=programa_id)
    return JsonResponse([{"id": s.id, "nombre": s.nombre} for s in subprogramas], safe=False)

@login_required
@rol_requerido("coordinador", "administrador")
def cargar_jardines(request):
    programa_id = request.GET.get("programa_id")
    jardines = Jardin.objects.filter(programa_id=programa_id)
    return JsonResponse([{"id": j.id, "nombre": j.nombre} for j in jardines], safe=False)

@login_required
@rol_requerido("coordinador", "administrador")
def validar_docente_turno(request):
    docente_id = request.GET.get("docente_id")
    turno = request.GET.get("turno")
    sala_id = request.GET.get("sala_id")
    qs = Sala.objects.filter(docentes__id=docente_id, turno=turno)
    if sala_id:
        qs = qs.exclude(id=sala_id)
    return JsonResponse({"conflicto": qs.exists()})

@login_required
def subprogramas_por_programa(request):
    programa_id = request.GET.get("programa_id")
    subprogramas = []
    if programa_id:
        subprogramas = list(Subprograma.objects.filter(programa_id=programa_id).values("id", "nombre"))
    return JsonResponse(subprogramas, safe=False)

# ... (vistas de asistencia docente a continuación) ...

@login_required
@rol_requerido("coordinador", "administrador")
def lista_jardines_asistencia(request):
    """
    Lista de jardines para que el coordinador seleccione uno y cargue asistencia docente.
    Si es coordinador, solo ve los jardines de sus programas asignados.
    """
    jardines = Jardin.objects.all().select_related("programa").order_by("nombre")
    
    return render(request, "jardines/asistencia_docente_lista.html", {
        "jardines": jardines
    })

@login_required
@rol_requerido("coordinador", "administrador")
def cargar_asistencia_docente(request, jardin_id):
    """
    Carga masiva de asistencia para los docentes de un jardín.
    Muestra una fila por (docente, turno) para que cada turno tenga su propio estado.
    """
    jardin = get_object_or_404(Jardin, id=jardin_id)

    # 🔒 Validación de propiedad
    if not request.user.es_admin() and jardin.programa not in request.user.programas_asignados.all():
        raise PermissionDenied("No tiene permiso para cargar asistencia en este espacio.")

    fecha_str = request.GET.get("fecha", str(timezone.localtime(timezone.now()).date()))
    fecha = date.fromisoformat(fecha_str)

    # Obtener todos los pares únicos (docente, turno) asignados a salas de este jardín
    from .models import LicenciaDocente
    salas = Sala.objects.filter(jardin=jardin).prefetch_related("docentes")
    
    # Construir lista de (docente, turno) únicos ordenada por docente y turno
    pares = {}  # { (docente_id, turno): docente_obj }
    for sala in salas:
        for docente in sala.docentes.filter(rol__in=["docente", "auxiliar"]):
            key = (docente.id, sala.turno)
            if key not in pares:
                pares[key] = (docente, sala.turno)
    pares_lista = sorted(pares.values(), key=lambda x: (x[0].last_name, x[0].first_name, x[1] or ""))

    if request.method == "POST":
        for docente, turno in pares_lista:
            campo = f"estado_{docente.id}_{turno}"
            obs_campo = f"obs_{docente.id}_{turno}"
            estado = request.POST.get(campo)
            observaciones = request.POST.get(obs_campo, "")
            
            # Forzar 'L' si tiene licencia activa para este turno
            licencia = LicenciaDocente.obtener_licencia_activa(docente, fecha, turno=turno)
            if licencia:
                estado = 'L'
                observaciones = f"Ausente por licencia ({licencia.get_tipo_licencia_display()})."
            
            if estado:
                AsistenciaDocente.objects.update_or_create(
                    docente=docente,
                    jardin=jardin,
                    turno=turno,
                    fecha=fecha,
                    defaults={
                        "estado": estado,
                        "observaciones": observaciones,
                        "registrado_por": request.user
                    }
                )
        
        messages.success(request, f"Asistencia docente del {fecha.strftime('%d/%m/%Y')} guardada correctamente.")
        return redirect("jardines:cargar_asistencia_docente", jardin_id=jardin.id)

    # Buscar asistencias ya cargadas para esta fecha
    asistencias_existentes = AsistenciaDocente.objects.filter(jardin=jardin, fecha=fecha)
    asistencias_dict = {(a.docente_id, a.turno): a for a in asistencias_existentes}

    docente_data = []
    for docente, turno in pares_lista:
        lic = LicenciaDocente.obtener_licencia_activa(docente, fecha, turno=turno)
        docente_data.append({
            "docente": docente,
            "turno": turno,
            "turno_label": dict(AsistenciaDocente.TURNO_CHOICES).get(turno, turno or "—"),
            "asistencia": asistencias_dict.get((docente.id, turno)),
            "licencia": lic,
        })

    return render(request, "jardines/asistencia_docente_form.html", {
        "jardin": jardin,
        "fecha": fecha_str,
        "docente_data": docente_data,
        "ESTADO_CHOICES": AsistenciaDocente.ESTADO_CHOICES
    })

@login_required
@rol_requerido("coordinador", "administrador")
def historial_asistencia_docente(request):
    """
    Historial general de asistencia docente con filtros avanzados.
    """
    user = request.user
    fecha_exacta = request.GET.get("fecha")
    mes = request.GET.get("mes")
    anio = request.GET.get("anio")
    jardin_id = request.GET.get("jardin")
    turno = request.GET.get("turno")
    sala_id = request.GET.get("sala")
    q = request.GET.get("q", "").strip()
    estado_jornada = request.GET.get("estado_jornada")

    # 🔒 Base queryset filtrado por programas del coordinador
    if user.es_admin():
        asistencias = AsistenciaDocente.objects.select_related("docente", "jardin", "registrado_por")
        jardines = Jardin.objects.all().order_by("nombre")
        salas = Sala.objects.all().order_by("nombre")
    else:
        programas = user.programas_asignados.all()
        asistencias = AsistenciaDocente.objects.filter(
            jardin__programa__in=programas
        ).select_related("docente", "jardin", "registrado_por")
        jardines = Jardin.objects.filter(programa__in=programas).order_by("nombre")
        salas = Sala.objects.filter(jardin__programa__in=programas).order_by("nombre")

    if fecha_exacta:
        asistencias = asistencias.filter(fecha=fecha_exacta)
    if mes:
        asistencias = asistencias.filter(fecha__month=mes)
    if anio:
        asistencias = asistencias.filter(fecha__year=anio)
    if jardin_id:
        asistencias = asistencias.filter(jardin_id=jardin_id)
        salas = salas.filter(jardin_id=jardin_id)
    if turno:
        asistencias = asistencias.filter(turno=turno)
    if sala_id:
        asistencias = asistencias.filter(docente__salas_asignadas__id=sala_id)
    if q:
        asistencias = asistencias.filter(
            Q(docente__first_name__icontains=q) |
            Q(docente__last_name__icontains=q) |
            Q(docente__dni__icontains=q) |
            Q(docente__username__icontains=q)
        )

    hoy = timezone.localtime(timezone.now()).date()
    if estado_jornada:
        if estado_jornada == "en_curso":
            asistencias = asistencias.filter(fichado=True, fichado_salida=False, fecha=hoy)
        elif estado_jornada == "finalizada":
            asistencias = asistencias.filter(fichado=True, fichado_salida=True)
        elif estado_jornada == "sin_salida":
            asistencias = asistencias.filter(fichado=True, fichado_salida=False, fecha__lt=hoy)
        elif estado_jornada == "sin_fichaje":
            asistencias = asistencias.filter(fichado=False, estado__in=['A', 'P', 'J', 'T', 'R'])
        elif estado_jornada == "licencia":
            asistencias = asistencias.filter(estado='L')
        elif estado_jornada == "actividad_especial":
            asistencias = asistencias.filter(estado='E')

    anios = list(range(2024, hoy.year + 1))
    meses = list(range(1, 13))

    return render(request, "jardines/asistencia_docente_historial.html", {
        "asistencias": asistencias.distinct().order_by("-fecha", "turno", "docente__last_name"),
        "jardines": jardines,
        "salas": salas,
        "meses": meses,
        "anios": anios,
        "fecha_sel": fecha_exacta,
        "mes_sel": mes,
        "anio_sel": anio,
        "jardin_sel": jardin_id,
        "turno_sel": turno,
        "sala_sel": sala_id,
        "q_sel": q,
        "estado_jornada_sel": estado_jornada,
        "TURNO_CHOICES": AsistenciaDocente.TURNO_CHOICES,
    })

@login_required
@rol_requerido("coordinador", "administrador")
def reporte_asistencia_docente_mensual(request):
    """
    Genera una matriz mensual de asistencia para todos los docentes de un jardín.
    Muestra una fila por (docente, turno) en lugar de una fila global por docente.
    """
    ahora = timezone.localtime(timezone.now()).date()
    mes = int(request.GET.get('mes', ahora.month))
    anio = int(request.GET.get('anio', ahora.year))
    jardin_id = request.GET.get('jardin')

    if not jardin_id:
        if request.user.es_admin():
            jardines = Jardin.objects.all().order_by("nombre")
        else:
            jardines = Jardin.objects.filter(
                programa__in=request.user.programas_asignados.all()
            ).order_by("nombre")

        return render(request, "jardines/reporte_docente_seleccion.html", {
            "jardines": jardines,
            "meses": range(1, 13),
            "anios": range(2024, ahora.year + 1),
            "mes_sel": mes,
            "anio_sel": anio
        })

    jardin = get_object_or_404(Jardin, id=jardin_id)

    # 🔒 Validación de propiedad
    if not request.user.es_admin() and jardin.programa not in request.user.programas_asignados.all():
        raise PermissionDenied("No tiene permiso para ver el reporte de este espacio.")
    ultimo_dia_mes = calendar.monthrange(anio, mes)[1]
    
    # Obtener pares únicos (docente, turno) asignados a salas de este jardín
    salas = Sala.objects.filter(jardin=jardin).prefetch_related("docentes")
    pares = {}
    for sala in salas:
        for docente in sala.docentes.filter(rol__in=["docente", "auxiliar"]):
            key = (docente.id, sala.turno)
            if key not in pares:
                pares[key] = (docente, sala.turno)
    pares_lista = sorted(pares.values(), key=lambda x: (x[0].last_name, x[0].first_name, x[1] or ""))

    # Obtener todas las asistencias del mes
    asistencias = AsistenciaDocente.objects.filter(
        jardin=jardin,
        fecha__year=anio,
        fecha__month=mes
    )
    
    # Mapear asistencias: { (docente_id, turno): { dia: estado } }
    mapa_asistencia = {}
    for a in asistencias:
        key = (a.docente_id, a.turno)
        if key not in mapa_asistencia:
            mapa_asistencia[key] = {}
        mapa_asistencia[key][a.fecha.day] = a.estado

    filas = []
    turno_dict = dict(AsistenciaDocente.TURNO_CHOICES)
    for docente, turno in pares_lista:
        celdas = []
        presentes = 0
        ausentes = 0
        licencias = 0
        exentos = 0
        key = (docente.id, turno)
        turno_str = turno_dict.get(turno, turno or "")
        
        for dia in range(1, ultimo_dia_mes + 1):
            estado = mapa_asistencia.get(key, {}).get(dia, '-')
            celdas.append(estado)
            if estado == 'P': presentes += 1
            if estado == 'A': ausentes += 1
            if estado == 'L': licencias += 1
            if estado == 'E': exentos += 1
        
        filas.append({
            'docente': docente,
            'turno': turno,
            'turno_str': turno_str,
            'celdas': celdas,
            'presentes': presentes,
            'ausentes': ausentes,
            'licencias': licencias,
            'exentos': exentos,
        })

    context = {
        'jardin': jardin,
        'mes': mes,
        'anio': anio,
        'mes_nombre': calendar.month_name[mes].capitalize(),
        'dias': range(1, ultimo_dia_mes + 1),
        'filas': filas,
    }

    if 'export' in request.GET:
        return exportar_asistencia_docente_csv(context)

    return render(request, "jardines/reporte_asistencia_docente_mensual.html", context)

def exportar_asistencia_docente_csv(context):
    """
    Exporta la matriz de asistencia docente a CSV.
    """
    response = HttpResponse(content_type='text/csv')
    filename = f"asistencia_docente_{context['jardin'].nombre}_{context['mes']}_{context['anio']}.csv"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    response.write('\ufeff'.encode('utf8'))
    
    writer = csv.writer(response, delimiter=';')
    
    # Encabezado
    header = ['Docente', 'Turno'] + [str(d) for d in context['dias']] + ['P', 'A', 'L', 'E']
    writer.writerow(header)
    
    for fila in context['filas']:
        row = (
            [f"{fila['docente'].last_name}, {fila['docente'].first_name}", fila['turno_str']]
            + fila['celdas']
            + [fila['presentes'], fila['ausentes'], fila.get('licencias', 0), fila.get('exentos', 0)]
        )
        writer.writerow(row)
        
    return response

@login_required
@rol_requerido("coordinador", "administrador")
def resumen_actividad_docente(request):
    """
    Muestra un resumen de los docentes que iniciaron sesión y sus últimas acciones.
    """
    from users.models import AccionAuditoria
    from django.db.models import Max, OuterRef, Subquery
    
    fecha_str = request.GET.get("fecha")
    if fecha_str:
        try:
            hoy = date.fromisoformat(fecha_str)
        except (ValueError, TypeError):
            hoy = timezone.localtime(timezone.now()).date()
    else:
        hoy = timezone.localtime(timezone.now()).date()
        
    user = request.user

    # Base de docentes y auxiliares a supervisar
    if user.es_admin() or not user.programas_asignados.exists():
        docentes = Usuario.objects.filter(rol__in=["docente", "auxiliar"])
    else:
        programas = user.programas_asignados.all()
        docentes_salas = Usuario.objects.filter(
            rol__in=["docente", "auxiliar"],
            salas_asignadas__jardin__programa__in=programas
        )
        docentes_asistencia_hoy = Usuario.objects.filter(
            rol__in=["docente", "auxiliar"],
            asistencias_docente__fecha=hoy,
            asistencias_docente__jardin__programa__in=programas
        )
        docentes = (docentes_salas | docentes_asistencia_hoy).distinct()

    docentes = docentes.order_by("last_name")

    # Obtener asistencias de hoy
    asistencias_hoy = AsistenciaDocente.objects.filter(fecha=hoy).select_related("jardin")
    asistencias_dict = {}
    for a in asistencias_hoy:
        if a.docente_id not in asistencias_dict:
            asistencias_dict[a.docente_id] = []
        asistencias_dict[a.docente_id].append(a)

    # Buscar todas las acciones de hoy para los docentes filtrados
    acciones_hoy = AccionAuditoria.objects.filter(
        fecha__date=hoy,
        usuario__in=docentes
    ).select_related("usuario").order_by("-fecha")
    
    acciones_por_usuario = {}
    for acc in acciones_hoy:
        if acc.usuario_id not in acciones_por_usuario:
            acciones_por_usuario[acc.usuario_id] = []
        acciones_por_usuario[acc.usuario_id].append(acc)

    resumen = []
    for d in docentes:
        asists = asistencias_dict.get(d.id, [])
        accs = acciones_por_usuario.get(d.id, [])
        
        # Ordenar asistencias por turno ('mañana', 'tarde')
        def sort_key_asist(a):
            t = a.turno or ""
            if t == "mañana":
                return (0, a.id)
            if t == "tarde":
                return (1, a.id)
            return (2, a.id)
        
        asists_ordenadas = sorted(asists, key=sort_key_asist)

        # Construir desglose detallado de todos los turnos del día
        turnos_detalle = []
        for a in asists_ordenadas:
            lat_a = float(a.latitude) if (a.fichado and a.latitude) else None
            lon_a = float(a.longitude) if (a.fichado and a.longitude) else None
            turnos_detalle.append({
                "id": a.id,
                "turno": a.turno,
                "turno_display": a.get_turno_display() if a.turno else "General",
                "jardin_nombre": a.jardin.nombre if a.jardin else "General",
                "fichado": a.fichado,
                "fichado_salida": a.fichado_salida,
                "hora_ingreso": a.hora_ingreso if a.fichado else None,
                "hora_salida": a.hora_salida if a.fichado_salida else None,
                "horas_trabajadas": a.horas_trabajadas_str,
                "estado_jornada": a.estado_jornada,
                "estado": a.estado,
                "ip": a.ip_address if a.fichado else None,
                "lat": lat_a,
                "lon": lon_a,
                "observaciones": a.observaciones or "",
                "asistencia": a,
            })
        
        # Determine if they checked in (fichó) today
        fichado = any(a.fichado for a in asists) if asists else False
        fichado_salida = any(a.fichado_salida for a in asists) if asists else False
        
        # Get coordinates if fichado
        lat = None
        lon = None
        for td in turnos_detalle:
            if td["lat"] and td["lon"]:
                lat = td["lat"]
                lon = td["lon"]
                break
        
        primer_asist = asists_ordenadas[0] if asists_ordenadas else None

        resumen.append({
            "docente": d,
            "asistencias": asists_ordenadas,
            "turnos_detalle": turnos_detalle,
            "tiene_doble_turno": len(turnos_detalle) > 1,
            "fichado": fichado,
            "fichado_salida": fichado_salida,
            "inicio_sesion": primer_asist.hora_ingreso if (primer_asist and primer_asist.hora_ingreso) else None,
            "hora_salida": primer_asist.hora_salida if (primer_asist and primer_asist.hora_salida) else None,
            "horas_trabajadas": primer_asist.horas_trabajadas_str if primer_asist else "—",
            "ip": primer_asist.ip_address if primer_asist else None,
            "ultima_accion": accs[0] if accs else None,
            "total_acciones": len(accs),
            "activo": len(asists) > 0 or len(accs) > 0,
            "lat": lat,
            "lon": lon,
        })

    activos_count = sum(1 for item in resumen if item['activo'])

    return render(request, "jardines/resumen_actividad_docente.html", {
        "resumen": resumen,
        "fecha": hoy,
        "activos_count": activos_count,
    })

from django.views.decorators.http import require_POST
from django.utils import timezone
from decimal import Decimal, InvalidOperation

@login_required
@rol_requerido("docente", "auxiliar")
@require_POST
def registrar_asistencia_docente(request):
    latitude_str = request.POST.get("latitude")
    longitude_str = request.POST.get("longitude")
    tipo = request.POST.get("tipo")  # 'ingreso' or 'salida'
    
    if not latitude_str or not longitude_str:
        return JsonResponse({"status": "error", "message": "Las coordenadas de ubicación son requeridas al fichar."}, status=400)
        
    try:
        latitude = Decimal(latitude_str).quantize(Decimal('.000001'))
        longitude = Decimal(longitude_str).quantize(Decimal('.000001'))
    except (InvalidOperation, ValueError, TypeError):
        return JsonResponse({"status": "error", "message": "Formato de coordenadas no válido."}, status=400)
        
    ahora_local = timezone.localtime(timezone.now())
    hoy = ahora_local.date()
    hora = ahora_local.time()
    ip = request.META.get('REMOTE_ADDR')
    from .models import AsistenciaDocente, inicializar_asistencia_diaria, LicenciaDocente, _turno_por_hora, Jardin
    from users.models import AccionAuditoria
    from django.core.exceptions import ValidationError
    
    # Verificar licencia activa para el turno que está fichando
    turno_fichaje = request.POST.get("turno") or _turno_por_hora(hora)
    licencia = LicenciaDocente.obtener_licencia_activa(request.user, hoy, turno=turno_fichaje)
    if licencia:
        return JsonResponse({
            "status": "error",
            "message": f"No puede registrar asistencia porque se encuentra de licencia ({licencia.get_tipo_licencia_display()})."
        }, status=400)

    # Turno, asistencia_id, jardin_id
    asistencia_id = request.POST.get("asistencia_id")
    jardin_id = request.POST.get("jardin_id")
    turno = request.POST.get("turno") or _turno_por_hora(hora)

    # Inicializar todos los registros del día si aún no existen
    inicializar_asistencia_diaria(request.user, request)

    # Buscar los registros del turno / espacio que está fichando
    asist_qs = AsistenciaDocente.objects.filter(docente=request.user, fecha=hoy)
    if asistencia_id:
        asist_qs = asist_qs.filter(id=asistencia_id)
    elif jardin_id and turno:
        asist_qs = asist_qs.filter(jardin_id=jardin_id, turno=turno)
    else:
        asist_qs = asist_qs.filter(turno=turno)

    if not asist_qs.exists():
        target_jardin = None
        if request.user.salas_asignadas.exists():
            if jardin_id:
                sala_jardin = request.user.salas_asignadas.filter(jardin_id=jardin_id).first()
                if sala_jardin:
                    target_jardin = sala_jardin.jardin
            if not target_jardin:
                sala_turno = request.user.salas_asignadas.filter(turno=turno).first()
                if sala_turno:
                    target_jardin = sala_turno.jardin
                else:
                    sala_cualquiera = request.user.salas_asignadas.first()
                    if sala_cualquiera:
                        target_jardin = sala_cualquiera.jardin

        if target_jardin:
            asist, _ = AsistenciaDocente.objects.get_or_create(
                docente=request.user,
                jardin=target_jardin,
                turno=turno,
                fecha=hoy,
                defaults={
                    "hora_ingreso": None,
                    "ip_address": ip,
                    "fuera_de_jornada": False,
                    "estado": "A",
                    "fichado": False,
                    "observaciones": "Registro inicializado al fichar."
                }
            )
            asist_qs = AsistenciaDocente.objects.filter(id=asist.id)
        else:
            return JsonResponse({
                "status": "error",
                "message": "No tenés salas asignadas para registrar asistencia. Por favor, contactáte con tu coordinador."
            }, status=400)

    asist_ref = asist_qs.first()
    
    # Auto-detectar tipo si no fue especificado explícitamente
    if not tipo:
        if not asist_ref.fichado:
            tipo = "ingreso"
        elif not asist_ref.fichado_salida:
            tipo = "salida"
        else:
            turno_disp = asist_ref.get_turno_display() if asist_ref.turno else turno
            return JsonResponse({
                "status": "already",
                "message": f"Ya registraste tu ingreso y salida del turno {turno_disp} hoy."
            })

    if tipo == "ingreso":
        # Si ya fue fichado ingreso
        if not asist_qs.filter(fichado=False).exists():
            turno_disp = asist_ref.get_turno_display() if asist_ref.turno else turno
            return JsonResponse({
                "status": "already",
                "message": f"Ya registraste tu ingreso del turno {turno_disp} hoy."
            })

        try:
            for asist in asist_qs.filter(fichado=False):
                asist.fichado = True
                asist.estado = 'P'
                asist.hora_ingreso = hora
                asist.ip_address = ip
                asist.latitude = latitude
                asist.longitude = longitude
                turno_disp = asist.get_turno_display() if asist.turno else turno
                asist.observaciones = f"Ingreso registrado por el docente el {hoy.strftime('%d/%m/%Y')} a las {hora.strftime('%H:%M')} hs. (turno {turno_disp})."
                asist.save()
        except ValidationError as e:
            error_msg = ", ".join(e.messages) if hasattr(e, "messages") else str(e)
            return JsonResponse({"status": "error", "message": f"Error de validación: {error_msg}"}, status=400)
            
        AccionAuditoria.objects.create(
            usuario=request.user,
            accion="modificacion",
            modelo="AsistenciaDocente",
            objeto_id=request.user.id,
            descripcion=f"Fichó ingreso de asistencia ({turno}) el {hoy.strftime('%d/%m/%Y')} a las {hora.strftime('%H:%M:%S')} hs. IP: {ip}"
        )
        
        turno_final = asist_ref.get_turno_display() if (asist_ref and asist_ref.turno) else turno
        return JsonResponse({
            "status": "success",
            "tipo": "ingreso",
            "message": f"Ingreso del turno {turno_final} registrado correctamente.",
            "turno": turno
        })

    elif tipo == "salida":
        # Verificar que haya registrado ingreso previo
        if asist_qs.filter(fichado=False).exists():
            return JsonResponse({
                "status": "error",
                "message": "No podés fichar salida sin haber registrado el ingreso previamente."
            }, status=400)

        # Si ya fue fichada la salida
        if not asist_qs.filter(fichado_salida=False).exists():
            turno_disp = asist_ref.get_turno_display() if asist_ref.turno else turno
            return JsonResponse({
                "status": "already",
                "message": f"Ya registraste tu salida del turno {turno_disp} hoy."
            })

        try:
            for asist in asist_qs.filter(fichado_salida=False):
                asist.fichado_salida = True
                asist.hora_salida = hora
                asist.ip_address_salida = ip
                asist.latitude_salida = latitude
                asist.longitude_salida = longitude
                turno_disp = asist.get_turno_display() if asist.turno else turno
                obs_prev = asist.observaciones or ""
                asist.observaciones = (obs_prev + f" | Salida registrada el {hoy.strftime('%d/%m/%Y')} a las {hora.strftime('%H:%M')} hs.").strip(" | ")
                asist.save()
        except ValidationError as e:
            error_msg = ", ".join(e.messages) if hasattr(e, "messages") else str(e)
            return JsonResponse({"status": "error", "message": f"Error de validación: {error_msg}"}, status=400)
            
        AccionAuditoria.objects.create(
            usuario=request.user,
            accion="modificacion",
            modelo="AsistenciaDocente",
            objeto_id=request.user.id,
            descripcion=f"Fichó salida de asistencia ({turno}) el {hoy.strftime('%d/%m/%Y')} a las {hora.strftime('%H:%M:%S')} hs. IP: {ip}"
        )
        
        turno_final = asist_ref.get_turno_display() if (asist_ref and asist_ref.turno) else turno
        asist_ref.refresh_from_db()
        horas_trabajadas = asist_ref.horas_trabajadas_str
        return JsonResponse({
            "status": "success",
            "tipo": "salida",
            "message": f"Salida del turno {turno_final} registrada correctamente. Horas trabajadas: {horas_trabajadas}.",
            "turno": turno,
            "horas_trabajadas": horas_trabajadas
        })

    else:
        return JsonResponse({"status": "error", "message": "Tipo de fichada no válido."}, status=400)

import urllib.request
import urllib.error
from django.core.cache import cache
from django.views.decorators.clickjacking import xframe_options_exempt
from django.utils.decorators import method_decorator
from django.db.models import Q, Count
from datetime import date
import calendar
from alumnos.models import Asistencia

def obtener_costos_docentes_api(target_date):
    """
    Obtiene los costos de docentes y auxiliares por DNI para un mes/año dado.
    Prioriza la API de Google Apps Script (Google Sheets).
    Fallback a último mes cargado por DNI y fallback a archivo Excel local si la API no está configurada o falla.
    """
    import ssl
    costos = {}
    costos_fallback_ultimo_mes = {}
    from django.conf import settings

    # Crear contexto SSL permisivo por si el servidor de prod no tiene certifi / CA estricto
    ssl_ctx = ssl.create_default_context()
    ssl_ctx.check_hostname = False
    ssl_ctx.verify_mode = ssl.CERT_NONE

    meses_es = {
        1: 'enero', 2: 'febrero', 3: 'marzo', 4: 'abril',
        5: 'mayo', 6: 'junio', 7: 'julio', 8: 'agosto',
        9: 'septiembre', 10: 'octubre', 11: 'noviembre', 12: 'diciembre'
    }
    
    nombre_mes = meses_es.get(target_date.month, '')
    mes_abrev = nombre_mes[:3] if nombre_mes else ''
    year_str = str(target_date.year)
    year_short = year_str[-2:]
    
    candidatos_mes = {
        nombre_mes,
        mes_abrev,
        f"{mes_abrev}-{year_short}",
        f"{nombre_mes}-{year_short}",
        f"{nombre_mes} {year_short}",
        f"{nombre_mes}-{year_str}",
        f"{target_date.month:02d}-{year_short}",
        f"{target_date.month}-{year_short}"
    }
    if target_date.month == 9:
        candidatos_mes.add(f"sept-{year_short}")
        candidatos_mes.add("sept")

    api_urls = []
    url1 = getattr(settings, 'GOOGLE_APPS_SCRIPT_COSTOS_URL', '')
    url2 = getattr(settings, 'GOOGLE_APPS_SCRIPT_COSTOS_URL_2', '')
    
    if url1: api_urls.append((url1, "sheet1_rul"))
    if url2: api_urls.append((url2, "sheet2_rrhh"))

    import requests
    import urllib3
    urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

    for api_url, tag in api_urls:
        cache_key = f"costos_google_sheets_rows_{tag}"
        cached_data = cache.get(cache_key)
        
        data_rows = None
        if cached_data:
            data_rows = cached_data
        else:
            try:
                resp = requests.get(
                    api_url,
                    headers={'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64)'},
                    timeout=3.0,
                    verify=False
                )
                if resp.status_code == 200:
                    payload = resp.json()
                    if isinstance(payload, dict) and "data" in payload:
                        data_rows = payload["data"]
                    elif isinstance(payload, list):
                        data_rows = payload
                    if data_rows:
                        cache.set(cache_key, data_rows, 86400)  # Caché persistente de 24 horas
            except Exception:
                data_rows = None

        if data_rows:
            for item in data_rows:
                try:
                    dni_raw = str(item.get("dni", "")).strip().replace(".", "").replace(" ", "")
                    if not dni_raw:
                        continue
                    dni = "".join(filter(str.isdigit, dni_raw))
                    if not dni:
                        continue
                    
                    mes_item = str(item.get("mes", "")).strip().lower()
                    costo_val = float(item.get("costo", 0) or 0)
                    while 0 < costo_val < 100000:
                        costo_val *= 1000  # Convertir montos expresados en miles o millones (ej: 2.566 -> 2566000 pesos)

                    dnis_to_add = [dni]
                    if len(dni) == 11 and dni[:2] in ('20', '27', '23', '24', '25', '26'):
                        dnis_to_add.append(dni[2:10])

                    if costo_val > 0:
                        for d_key in dnis_to_add:
                            if d_key not in costos_fallback_ultimo_mes or tag == "sheet2_rrhh":
                                costos_fallback_ultimo_mes[d_key] = costo_val

                    if mes_item in candidatos_mes or (nombre_mes == mes_item) or ((nombre_mes in mes_item or mes_abrev in mes_item) and year_short in mes_item):
                        for d_key in dnis_to_add:
                            if d_key not in costos or (tag == "sheet2_rrhh" and costo_val > 0):
                                costos[d_key] = costo_val

                except (ValueError, TypeError):
                    pass

    for dni, costo_f in costos_fallback_ultimo_mes.items():
        if dni not in costos or costos[dni] == 0:
            costos[dni] = costo_f

    # --- SUPPLEMENT / FALLBACK CON EXCEL LOCAL ---
    # Cargar archivo Excel (con caché de 10 minutos para respuesta instantánea)
    excel_cache_key = f"costos_excel_local_parsed_{target_date.year}_{target_date.month}"
    costos_excel = cache.get(excel_cache_key)

    if costos_excel is None:
        costos_excel = {}
        try:
            excel_candidates = [
                os.path.join(settings.BASE_DIR, "Locaciones 02. Secretaria de Gestión y Participación Ciudadana.xlsx"),
                os.path.join(settings.BASE_DIR, "data", "Locaciones 02. Secretaria de Gestión y Participación Ciudadana.xlsx")
            ]
            excel_path = next((p for p in excel_candidates if os.path.exists(p)), None)
            
            if excel_path:
                import openpyxl
                wb = openpyxl.load_workbook(excel_path, data_only=True, read_only=True)
                sheet = wb['Locaciones']
                header_row = [cell for cell in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))]
                
                target_col_idx = None
                for idx, val in enumerate(header_row):
                    if hasattr(val, 'year') and hasattr(val, 'month'):
                        if val.year == target_date.year and val.month == target_date.month:
                            target_col_idx = idx
                            break

                for row in sheet.iter_rows(min_row=2, values_only=True):
                    try:
                        val = row[2]
                        if not val: continue
                        dni_clean = "".join(filter(str.isdigit, str(val)))
                        if not dni_clean: continue

                        c_val = 0
                        if target_col_idx is not None and target_col_idx < len(row):
                            try: c_val = float(row[target_col_idx] or 0)
                            except: pass

                        if c_val == 0:
                            for cell_v in reversed(row[8:]):
                                try:
                                    num = float(cell_v or 0)
                                    if num > 0:
                                        c_val = num
                                        break
                                except: pass

                        while 0 < c_val < 100000:
                            c_val *= 1000

                        if c_val > 0:
                            dnis_list = [dni_clean]
                            if len(dni_clean) == 11 and dni_clean[:2] in ('20', '27', '23', '24', '25', '26'):
                                dnis_list.append(dni_clean[2:10])

                            for d_key in dnis_list:
                                if d_key not in costos_excel:
                                    costos_excel[d_key] = c_val
                    except Exception:
                        pass
                wb.close()
        except Exception:
            pass

        cache.set(excel_cache_key, costos_excel, 600)

    # Rellenar cualquier DNI no encontrado en Google Sheets con los datos de la planilla Excel
    for d_key, c_val in costos_excel.items():
        if d_key not in costos or costos[d_key] == 0:
            costos[d_key] = c_val

    return costos


@method_decorator(xframe_options_exempt, name='dispatch')
class BaseDashboardProgramaView(TemplateView):
    template_name = "jardines/dashboard_espacios_ludicos.html"
    programa_id = None
    programa_nombre = None
    programa_titulo = ""
    programa_subtitulo = ""

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Obtener Objeto Programa
        programa_obj = None
        if self.programa_id:
            programa_obj = Programa.objects.filter(pk=self.programa_id).first()
        elif self.programa_nombre:
            programa_obj = Programa.objects.filter(nombre=self.programa_nombre).first()
        
        if not programa_obj and self.programa_nombre:
            programa_obj = Programa.objects.filter(nombre__icontains=self.programa_nombre).first()

        # Filtros
        subprograma_id = self.request.GET.get('subprograma')
        zona_filtro = self.request.GET.get('zona')
        fecha_filtro = self.request.GET.get('fecha')
        
        # Procesar Fecha
        hoy = date.today()
        meses_es = {
            1: 'ENERO', 2: 'FEBRERO', 3: 'MARZO', 4: 'ABRIL', 
            5: 'MAYO', 6: 'JUNIO', 7: 'JULIO', 8: 'AGOSTO', 
            9: 'SEPTIEMBRE', 10: 'OCTUBRE', 11: 'NOVIEMBRE', 12: 'DICIEMBRE'
        }
        
        if fecha_filtro:
            try:
                filtro_year, filtro_month = map(int, fecha_filtro.split('-'))
                target_date = date(filtro_year, filtro_month, 1)
                mes_label = f"{meses_es[filtro_month]} {filtro_year}"
            except ValueError:
                target_date = hoy.replace(day=1)
                mes_label = f"{meses_es[hoy.month]} {hoy.year}"
        else:
            target_date = hoy.replace(day=1)
            mes_label = f"{meses_es[hoy.month]} {hoy.year}"
            
        start_of_selected_month = target_date
        end_of_selected_month = target_date.replace(day=calendar.monthrange(target_date.year, target_date.month)[1])
        
        # --- COST PARSING (API Google Sheets / Excel Fallback) ---
        costos_docentes = obtener_costos_docentes_api(target_date)
        # ---------------------------
        
        q_activo_mes = Q(
            Q(fecha_ingreso__lte=end_of_selected_month) | Q(alumno__asistencias__fecha__lte=end_of_selected_month)
        ) & ~Q(
            activo=False,
            fecha_baja__lt=start_of_selected_month
        )
        
        q_baja_mes = Q(
            activo=False, 
            fecha_baja__gte=start_of_selected_month, 
            fecha_baja__lte=end_of_selected_month
        )
        
        # Base querysets
        if programa_obj:
            jardines = Jardin.objects.filter(programa=programa_obj)
        else:
            jardines = Jardin.objects.none()
            
        if subprograma_id:
            jardines = jardines.filter(subprograma_id=subprograma_id)
        if zona_filtro:
            jardines = jardines.filter(sector=zona_filtro)
            
        jardines_ids = jardines.values_list('id', flat=True)
        
        # Alcance: Alumnos asignados a salas de estos jardines
        asignaciones = AsignacionSala.objects.filter(sala__jardin_id__in=jardines_ids)
        
        q_inscriptos = Q(fecha_ingreso__lte=end_of_selected_month) | Q(alumno__asistencias__fecha__lte=end_of_selected_month)
        total_inscriptos = asignaciones.filter(q_inscriptos).values('alumno').distinct().count()
        activos = asignaciones.filter(q_activo_mes).values('alumno').distinct().count()
        bajas = asignaciones.filter(q_baja_mes).values('alumno').distinct().count()
        
        cantidad_espacios = jardines.count()
        
        # Docentes
        cantidad_docentes = Usuario.objects.filter(
            rol__in=["docente", "auxiliar"],
            salas_asignadas__jardin_id__in=jardines_ids
        ).distinct().count()
        
        # Gráfico Mensual (Activos por mes, últimos 6 meses)
        hoy = date.today()
        meses = []
        cantidades = []
        crecimiento = []
        
        anterior = 0
        for i in range(5, -1, -1):
            target_month = hoy.month - i
            target_year = hoy.year
            while target_month <= 0:
                target_month += 12
                target_year -= 1
            
            t_date = date(target_year, target_month, 1)
            start_of_month = t_date
            end_of_month = t_date.replace(day=calendar.monthrange(target_year, target_month)[1])
            
            activos_mes = asignaciones.filter(
                Q(fecha_ingreso__lte=end_of_month) | Q(alumno__asistencias__fecha__lte=end_of_month)
            ).exclude(
                activo=False,
                fecha_baja__lt=start_of_month
            ).values('alumno').distinct().count()
            
            meses.append(t_date.strftime("%b %Y"))
            cantidades.append(activos_mes)
            
            if i == 5:
                crecimiento.append(0)  # Primer mes no hay % con respecto al anterior
            else:
                if anterior > 0:
                    pct = ((activos_mes - anterior) / anterior) * 100
                else:
                    pct = 100 if activos_mes > 0 else 0
                crecimiento.append(round(pct, 1))
            anterior = activos_mes

        # Gráfico Torta (Activos por Sector/Zona)
        sectores_qs = asignaciones.filter(q_activo_mes).values('sala__jardin__sector').annotate(total=Count('alumno', distinct=True))
        pie_labels = []
        pie_data = []
        for s in sectores_qs:
            pie_labels.append(s['sala__jardin__sector'] or "Sin sector")
            pie_data.append(s['total'])
            
        # Datos para tabla de Espacios y Mapa
        espacios_dict = {}
        jardines_mapa = []
        
        filtros_jardin = {}
        if programa_obj:
            filtros_jardin["programa"] = programa_obj
        if subprograma_id:
            filtros_jardin["subprograma_id"] = subprograma_id
        if zona_filtro:
            filtros_jardin["sector"] = zona_filtro
            
        if programa_obj:
            jardines_obj = Jardin.objects.filter(**filtros_jardin).order_by('nombre')
        else:
            jardines_obj = Jardin.objects.none()
            
        from django.db.models import Prefetch
        salas_jardin_qs = Sala.objects.filter(jardin__in=jardines_obj).select_related('jardin').prefetch_related(
            'docentes',
            Prefetch('alumnos_asignados', queryset=AsignacionSala.objects.filter(q_activo_mes), to_attr='activos_list'),
            Prefetch('alumnos_asignados', queryset=AsignacionSala.objects.filter(q_baja_mes), to_attr='bajas_list')
        )
        
        # Agrupar salas por jardin_id
        salas_por_jardin = {}
        for s in salas_jardin_qs:
            j_id = s.jardin_id
            if j_id not in salas_por_jardin:
                salas_por_jardin[j_id] = []
            salas_por_jardin[j_id].append(s)
            
        # Calcular horas totales por docente
        docente_horas_totales = {}
        for s in salas_jardin_qs:
            horas = 0
            if s.horario_inicio and s.horario_fin:
                h_in = s.horario_inicio.hour + s.horario_inicio.minute / 60
                h_fi = s.horario_fin.hour + s.horario_fin.minute / 60
                horas = h_fi - h_in
                if horas <= 0: horas = 1
            else:
                horas = 1
                
            for d in s.docentes.all():
                dni_db = str(d.dni).strip().replace('.', '').replace(' ', '')
                if not dni_db:
                    dni_db = str(d.username).strip().replace('.', '').replace(' ', '')
                
                if dni_db not in docente_horas_totales:
                    docente_horas_totales[dni_db] = 0
                docente_horas_totales[dni_db] += horas
            
        for j in jardines_obj:
            espacios_dict[j.nombre] = {"turnos": {}, "total_activos": 0, "total_bajas": 0, "costo_total": 0}
            
            if j.coordenadas:
                try:
                    coord_clean = j.coordenadas.replace("'", "").replace('"', '')
                    lat, lon = map(float, coord_clean.split(','))
                    jardines_mapa.append({
                        "nombre": j.nombre,
                        "lat": lat,
                        "lon": lon,
                        "activos": 0,
                        "zona": j.sector or "Sin Zona",
                        "subprograma": j.subprograma.nombre if j.subprograma else "Sin Subprograma",
                    })
                except: pass
                
            salas_jardin = salas_por_jardin.get(j.id, [])
            
            for s in salas_jardin:
                if s.turno not in espacios_dict[j.nombre]["turnos"]:
                    espacios_dict[j.nombre]["turnos"][s.turno] = {"salas": [], "total_activos": 0, "total_bajas": 0, "costo_total": 0}
                    
                docentes = []
                costo_sala = 0
                for d in s.docentes.all():
                    dni_db = str(d.dni).strip().replace('.', '').replace(' ', '')
                    if not dni_db:
                        dni_db = str(d.username).strip().replace('.', '').replace(' ', '')
                        
                    horas_sala = 0
                    if s.horario_inicio and s.horario_fin:
                        h_in = s.horario_inicio.hour + s.horario_inicio.minute / 60
                        h_fi = s.horario_fin.hour + s.horario_fin.minute / 60
                        horas_sala = h_fi - h_in
                        if horas_sala <= 0: horas_sala = 1
                    else:
                        horas_sala = 1
                        
                    horas_totales = docente_horas_totales.get(dni_db, 1)
                    proporcion = (horas_sala / horas_totales) if horas_totales > 0 else 0
                    
                    c_total = costos_docentes.get(dni_db, 0)
                    c_proporcional = c_total * proporcion
                    costo_sala += c_proporcional
                    docentes.append({
                        "nombres": d.get_full_name() or d.username,
                        "costo": c_proporcional
                    })
                    
                activos_sala = len(set(a.alumno_id for a in s.activos_list))
                bajas_sala = len(set(a.alumno_id for a in s.bajas_list))
                total_sala = activos_sala + bajas_sala
                
                permanencia = (activos_sala / total_sala * 100) if total_sala > 0 else 0
                desercion = (bajas_sala / total_sala * 100) if total_sala > 0 else 0
                
                costo_x_alumno_sala = (costo_sala / activos_sala) if activos_sala > 0 else 0
                
                espacios_dict[j.nombre]["turnos"][s.turno]["salas"].append({
                    "nombre": s.nombre,
                    "activos": activos_sala,
                    "bajas": bajas_sala,
                    "permanencia": permanencia,
                    "desercion": desercion,
                    "docentes": docentes,
                    "costo_total": costo_sala,
                    "costo_x_alumno": costo_x_alumno_sala
                })
                espacios_dict[j.nombre]["turnos"][s.turno]["total_activos"] += activos_sala
                espacios_dict[j.nombre]["turnos"][s.turno]["total_bajas"] += bajas_sala
                espacios_dict[j.nombre]["turnos"][s.turno]["costo_total"] += costo_sala
                
                espacios_dict[j.nombre]["total_activos"] += activos_sala
                espacios_dict[j.nombre]["total_bajas"] += bajas_sala
                espacios_dict[j.nombre]["costo_total"] += costo_sala
                
                # Update map activos
                existente = next((m for m in jardines_mapa if m["nombre"] == j.nombre), None)
                if existente:
                    existente["activos"] += activos_sala
                    
        # Calcular porcentajes globales por Jardín y Turno
        for j_nombre, datos_jardin in espacios_dict.items():
            t_act = datos_jardin["total_activos"]
            t_baj = datos_jardin["total_bajas"]
            t_costo = datos_jardin["costo_total"]
            t_tot = t_act + t_baj
            datos_jardin["permanencia"] = (t_act / t_tot * 100) if t_tot > 0 else 0
            datos_jardin["desercion"] = (t_baj / t_tot * 100) if t_tot > 0 else 0
            datos_jardin["costo_x_alumno"] = (t_costo / t_act) if t_act > 0 else 0
            
            for t_nombre, datos_turno in datos_jardin["turnos"].items():
                tt_act = datos_turno["total_activos"]
                tt_baj = datos_turno["total_bajas"]
                tt_costo = datos_turno["costo_total"]
                tt_tot = tt_act + tt_baj
                datos_turno["permanencia"] = (tt_act / tt_tot * 100) if tt_tot > 0 else 0
                datos_turno["desercion"] = (tt_baj / tt_tot * 100) if tt_tot > 0 else 0
                datos_turno["costo_x_alumno"] = (tt_costo / tt_act) if tt_act > 0 else 0
            
        # Preparar reporte mensual (matriz de asistencias)
        asistencias_mes = Asistencia.objects.filter(
            sala__jardin_id__in=jardines_ids,
            fecha__gte=start_of_selected_month,
            fecha__lte=end_of_selected_month,
            estado__in=['P', 'T', 'R']
        ).values('sala__jardin__nombre', 'sala__turno', 'sala__nombre', 'fecha').annotate(total=Count('id'))
        
        dias_rango = list(range(1, end_of_selected_month.day + 1))
        
        matriz_asistencia = {}
        
        total_costo_global = 0
        total_activos_global = 0
        
        for j in jardines:
            e_dict = espacios_dict.get(j.nombre, {})
            matriz_asistencia[j.nombre] = {
                "total_activos": e_dict.get("total_activos", 0),
                "costo_total": e_dict.get("costo_total", 0),
                "costo_x_alumno": e_dict.get("costo_x_alumno", 0),
                "dias": {d: 0 for d in dias_rango},
                "turnos": {}
            }
            # Cargar estructura vacía desde espacios_dict
            if j.nombre in espacios_dict:
                total_costo_global += e_dict.get("costo_total", 0)
                total_activos_global += e_dict.get("total_activos", 0)
                
                for t_nombre, t_datos in espacios_dict[j.nombre]["turnos"].items():
                    matriz_asistencia[j.nombre]["turnos"][t_nombre] = {
                        "total_activos": t_datos.get("total_activos", 0),
                        "costo_total": t_datos.get("costo_total", 0),
                        "costo_x_alumno": t_datos.get("costo_x_alumno", 0),
                        "dias": {d: 0 for d in dias_rango},
                        "salas": {}
                    }
                    for s in t_datos["salas"]:
                        matriz_asistencia[j.nombre]["turnos"][t_nombre]["salas"][s["nombre"]] = {
                            "total_activos": s.get("activos", 0),
                            "costo_total": s.get("costo_total", 0),
                            "costo_x_alumno": s.get("costo_x_alumno", 0),
                            "dias": {d: 0 for d in dias_rango}
                        }
        
        for a in asistencias_mes:
            j_nombre = a['sala__jardin__nombre']
            t_nombre = a['sala__turno']
            s_nombre = a['sala__nombre']
            dia = a['fecha'].day
            if j_nombre in matriz_asistencia:
                matriz_asistencia[j_nombre]["dias"][dia] += a['total']
                if t_nombre in matriz_asistencia[j_nombre]["turnos"]:
                    matriz_asistencia[j_nombre]["turnos"][t_nombre]["dias"][dia] += a['total']
                    if s_nombre in matriz_asistencia[j_nombre]["turnos"][t_nombre]["salas"]:
                        matriz_asistencia[j_nombre]["turnos"][t_nombre]["salas"][s_nombre]["dias"][dia] += a['total']
                        
        # Calcular dias con asistencia
        for j_nombre, j_datos in matriz_asistencia.items():
            j_datos["dias_con_asist"] = sum(1 for d, t in j_datos["dias"].items() if t > 0)
            for t_nombre, t_datos in j_datos["turnos"].items():
                t_datos["dias_con_asist"] = sum(1 for d, t in t_datos["dias"].items() if t > 0)
                for s_nombre, s_datos in t_datos["salas"].items():
                    s_datos["dias_con_asist"] = sum(1 for d, t in s_datos["dias"].items() if t > 0)

        # Convert to list and sort
        reporte_mensual = [{"espacio": k, **v} for k, v in matriz_asistencia.items()]
        reporte_mensual.sort(key=lambda x: x["espacio"])

        subprog_qs = Subprograma.objects.filter(programa=programa_obj) if programa_obj else Subprograma.objects.none()

        context.update({
            "programa_titulo": self.programa_titulo or (programa_obj.nombre if programa_obj else ""),
            "programa_subtitulo": self.programa_subtitulo or (f"Dashboard exclusivo de {programa_obj.nombre}" if programa_obj else ""),
            "total_inscriptos": total_inscriptos,
            "activos": activos,
            "bajas": bajas,
            "cantidad_espacios": cantidad_espacios,
            "cantidad_docentes": cantidad_docentes,
            "costo_total_docentes": total_costo_global,
            "costo_x_alumno_global": (total_costo_global / total_activos_global) if total_activos_global > 0 else 0,
            "mapa_data_json": json.dumps(jardines_mapa),
            "grafico_labels_json": json.dumps(meses),
            "grafico_data_json": json.dumps(cantidades),
            "grafico_crecimiento_json": json.dumps(crecimiento),
            "pie_labels_json": json.dumps(pie_labels),
            "pie_data_json": json.dumps(pie_data),
            "espacios_dict": espacios_dict,
            "reporte_mensual": reporte_mensual,
            "dias_mes": dias_rango,
            "mes_nombre_reporte": mes_label,
            "subprogramas": subprog_qs,
            "subprograma_sel": int(subprograma_id) if subprograma_id and subprograma_id.isdigit() else '',
            "zona_sel": zona_filtro or '',
            "sectores": Jardin.SECTORES_CHOICES,
        })
        return context


class DashboardEspaciosLudicosView(BaseDashboardProgramaView):
    programa_id = 2
    programa_nombre = "Espacios lúdicos y de aprendizaje para la primera infancia"
    programa_titulo = "Espacio educativo para la primer infancia (jardines Maternales municipales)"
    programa_subtitulo = "Dashboard exclusivo para la primera infancia."


class DashboardAlfabetizacionView(BaseDashboardProgramaView):
    programa_id = 6
    programa_nombre = "Programa Municipal de Alfabetización y Programa Acompañamiento Educativo"
    programa_titulo = "Programa Municipal de Alfabetización y Acompañamiento Educativo"
    programa_subtitulo = "Dashboard de gestión del Programa Municipal de Alfabetización y Acompañamiento Educativo."


class DashboardCarpinteriaView(BaseDashboardProgramaView):
    programa_id = 7
    programa_nombre = "Escuela de Carpintería"
    programa_titulo = "Escuela de Carpintería"
    programa_subtitulo = "Dashboard de gestión de la Escuela de Carpintería."


class DashboardArtesPlasticasView(BaseDashboardProgramaView):
    programa_id = 8
    programa_nombre = "Escuela Municipal de Artes Plásticas Manuel Belgrano"
    programa_titulo = "Escuela Municipal de Artes Plásticas Manuel Belgrano"
    programa_subtitulo = "Dashboard de gestión de la Escuela Municipal de Artes Plásticas Manuel Belgrano."


class DashboardExpresionCulturalView(BaseDashboardProgramaView):
    programa_id = 9
    programa_nombre = "Expresión Cultural"
    programa_titulo = "Expresión Cultural"
    programa_subtitulo = "Dashboard de gestión del Programa Expresión Cultural."


@csrf_exempt
def api_dashboard_datos(request):
    """
    API en tiempo real que devuelve la estructura completa de datos del tablero:
    Métricas, evolución histórica, gráfico por zonas, mapa interactivo, tabla jerárquica de espacios y matriz de asistencias.
    Soporta CORS y filtros por subprograma, zona y período.
    """
    programa_param = request.GET.get('programa')
    
    views_map = {
        '2': DashboardEspaciosLudicosView,
        '6': DashboardAlfabetizacionView,
        '7': DashboardCarpinteriaView,
        '8': DashboardArtesPlasticasView,
        '9': DashboardExpresionCulturalView,
        'espacios-ludicos': DashboardEspaciosLudicosView,
        'alfabetizacion': DashboardAlfabetizacionView,
        'carpinteria': DashboardCarpinteriaView,
        'artes-plasticas': DashboardArtesPlasticasView,
        'expresion-cultural': DashboardExpresionCulturalView,
    }
    
    view_cls = views_map.get(str(programa_param), DashboardEspaciosLudicosView)
    view_inst = view_cls()
    view_inst.request = request

    ctx = view_inst.get_context_data()

    subprogramas_data = [{"id": s.id, "nombre": s.nombre} for s in ctx.get("subprogramas", [])]
    sectores_data = [{"val": s[0], "label": s[1]} for s in ctx.get("sectores", [])]

    data = {
        "programa_titulo": ctx.get("programa_titulo", ""),
        "programa_subtitulo": ctx.get("programa_subtitulo", ""),
        "total_inscriptos": ctx.get("total_inscriptos", 0),
        "activos": ctx.get("activos", 0),
        "bajas": ctx.get("bajas", 0),
        "cantidad_espacios": ctx.get("cantidad_espacios", 0),
        "cantidad_docentes": ctx.get("cantidad_docentes", 0),
        "costo_total_docentes": ctx.get("costo_total_docentes", 0),
        "costo_x_alumno_global": round(ctx.get("costo_x_alumno_global", 0)),
        "grafico_labels": json.loads(ctx.get("grafico_labels_json", "[]")),
        "grafico_data": json.loads(ctx.get("grafico_data_json", "[]")),
        "grafico_crecimiento": json.loads(ctx.get("grafico_crecimiento_json", "[]")),
        "pie_labels": json.loads(ctx.get("pie_labels_json", "[]")),
        "pie_data": json.loads(ctx.get("pie_data_json", "[]")),
        "mapa_data": json.loads(ctx.get("mapa_data_json", "[]")),
        "espacios_dict": ctx.get("espacios_dict", {}),
        "reporte_mensual": ctx.get("reporte_mensual", []),
        "dias_mes": ctx.get("dias_mes", []),
        "mes_nombre_reporte": ctx.get("mes_nombre_reporte", ""),
        "subprogramas": subprogramas_data,
        "subprograma_sel": ctx.get("subprograma_sel", ""),
        "zona_sel": ctx.get("zona_sel", ""),
        "sectores": sectores_data,
    }

    response = JsonResponse(data)
    response["Access-Control-Allow-Origin"] = "*"
    response["Access-Control-Allow-Headers"] = "*"
    return response


